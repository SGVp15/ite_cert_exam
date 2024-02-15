import re

from openpyxl import load_workbook

from Contact import Contact
from config import OUT_DIR
from utils.translit import replace_month_to_number
from .config import Columns, FILE_XLSX, PAGE_NAME


def get_contact_from_excel() -> list[Contact]:
    # Прочитать XLSX file -> [Contacts]
    filename = FILE_XLSX
    file_excel = load_workbook(filename=filename, data_only=True)
    sheet_names = file_excel.sheetnames
    contacts = []
    for row in range(2, file_excel[sheet_names[0]].max_row):
        contact = Contact()
        try:
            contact.email = clean_export_excel(read_excel(file_excel, column=Columns.email, row=row))
            if not re.search('@', contact.email):
                continue
            contact.number = int(clean_export_excel(read_excel(file_excel, column=Columns.number, row=row)))
            contact.abr_exam = clean_export_excel(read_excel(file_excel, column=Columns.abr_exam, row=row))
            contact.name_rus = clean_export_excel(read_excel(file_excel, column=Columns.name_rus, row=row))
            contact.name_eng = clean_export_excel(read_excel(file_excel, column=Columns.name_eng, row=row))
            contact.date_exam = clean_export_excel(read_excel(file_excel, column=Columns.date_exam, row=row))
            contact.template = contact.abr_exam + '.png'

        except Exception as e:
            print(e)
            continue
        # Создаем папки по курсам
        date_exam = f"{contact.date_exam}"
        date_exam = date_exam.replace(' ', '')
        date_exam = replace_month_to_number(date_exam)
        date_exam = str('.'.join(date_exam.split('.')[::-1]))
        contact.dir_name = date_exam

        certificate = 'Сертификат'

        file_out_png = f"{OUT_DIR}{date_exam}/{certificate}_{date_exam}_" \
                       f"{contact.name_rus}_{contact.number}_{contact.email}.png"

        file_out_png = file_out_png.replace(' ', '_')
        file_out_png = replace_month_to_number(file_out_png)
        contact.file_out_png = file_out_png

        contacts.append(contact)

    return contacts


def read_excel(excel, column, row):
    sheet_ranges = excel[PAGE_NAME]
    return str(sheet_ranges[f'{column}{row}'].value)


def clean_export_excel(s):
    s = s.replace(',', ', ')
    s = re.sub(r'\s{2,}', ' ', s)
    s = s.strip()
    if s in ('None', '#N/A'):
        s = ''
    return s
